import requests
import json
from openpyxl import load_workbook
from datetime import date


TBC_SHEET_NAMES = {'Summary', 'transactions_history'}
NB_API = 'https://nbg.gov.ge/gw/api/ct/monetarypolicy/currencies/en/json/?currencies={}&date={}'


def parse_tbs_statement(sheet):
    data = {}
    for row in sheet.values:
        if row[0] is not None:
            if row[2] == 'Income':
                format_date = f'{row[0].year}-{row[0].month}-{row[0].day}'
                if format_date in data.keys():
                    day_values = data[format_date]
                    currency = day_values.get(row[4])
                    day_values[row[4]] = currency + row[3] if currency else row[3]
                    data[format_date] = day_values
                else:
                    data[format_date] = {row[4]: row[3]}
        else:
            break
    return data


def exchange_rate(currency, date):
    response = requests.get(NB_API.format(currency, date))
    data = json.loads(response.content)
    rate = data[0]['currencies'][0]['rate']
    quantity = data[0]['currencies'][0]['quantity']
    val = rate / quantity
    return val


def calculate(file):
    wb = load_workbook(filename=file)
    sheet = wb.get_sheet_by_name('transactions_history')
    data = parse_tbs_statement(sheet)
    total = 0
    for day in data.keys():
        for currency, val in data[day].items():
            if currency != 'GEL':
                rate = exchange_rate(currency, day)
                val = val * rate
                total += val
            else:
                total += val
    return total


def check_sheets(sheet_names, bank_statement_sheet_names):
    for name in sheet_names:
        return True if name in bank_statement_sheet_names else False


def check_tbc(sheet):
    val = sheet['C3'].value
    return True if val[4:6] == 'TB' else False


def check_date_in_tbs(sheet):
    sheet_date = sheet['A3'].value
    current_date = date.today()
    check_year = sheet_date.year == current_date.year
    check_month = current_date.month - sheet_date.month == 1
    return check_year and check_month


def is_correct_statement(file, user):
    wb = load_workbook(filename=file)
    sheet_names = wb.get_sheet_names()

    # user bank support check
    if user.profile.bank == 'TBC':

        # check required sheets in the statement
        if not check_sheets(sheet_names, TBC_SHEET_NAMES):
            return False

        # check TB teg in account number
        sheet = wb.get_sheet_by_name('Summary')
        if not check_tbc(sheet):
            return False

        # check correct date of statement
        sheet = wb.get_sheet_by_name('transactions_history')
        if not check_date_in_tbs(sheet):
            return False

        return True

    else:
        return False



